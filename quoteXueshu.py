import openpyxl
import requests
import re
class quoteXueshu:
    xushuurl = 'http://xueshu.baidu.com/'                 #百度学术主页
    mainsearchurl = 'http://xueshu.baidu.com/s'           #搜索文献主页
    downmainurl = "http://xueshu.baidu.com/u/citation"  # 下载引用的主链接
    quote_format = 'enw'                                     #默认下载endnote格式

    namexlsxPath = ''                                        #论文名称的表格，第一列为名称
    savePath = ''                                            #保存的路径

    paperNameList = ''                                       #paper名称列表

    timesleep = 1                                            #默认一秒间隔
    #头信息
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36",
        "Host": "xueshu.baidu.com",
        "Upgrade-Insecure-Requests": "1",
        "Connection": "keep-alive",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
        "Accept-Encoding": "gzip,deflate",
        "Accept-Language": "zh-CN,zh;q=0.9",
        }
    #搜索参数
    searchdata = {
        "wd": '',
        "bs": '',
        "tn": "SE_baiduxueshu_c1gjeupa",
        "ie": "utf-8",
        "f": "8",
        "rsv_sug2": "1",
        "sc_f_para": "sc_tasktype={firstSimpleSearch}",
        "rsv_n": "2",
        "cl":"3",
        "rsv_bp":"1",
        "rsv_spt":"3"
    }
    #下载参数
    endnote_para = {
        "url": '',
        "sign": '',
        "diversion": '',
        "t": quote_format
    }
    def __init__(self,namexlsxPath,savePath):
        self.namexlsxPath = namexlsxPath
        self.savePath = savePath
        # print(savePath)

    def getMainHtml(self,paperName):
        #去除特殊格式
        paperName = paperName.replace("/", "_")
        paperName = paperName.replace(":", "_")
        paperName = paperName.replace("\n", "_")
        self.searchdata['wd'] = paperName
        self.searchdata['bs'] = paperName

        html = requests.get(self.mainsearchurl,params = self.searchdata,headers = self.headers, )
        # html = requests.get('http://xueshu.baidu.com/s?tn=SE_baiduxueshu_c1gjeupa&wd=%E5%90%91%E5%B1%B1%E5%9E%83%E5%9C%BE%E5%A1%AB%E5%9F%8B%E5%9C%BA-%E5%B0%BE%E7%9F%BF%E6%B7%B7%E5%90%88%E5%B8%A6%E7%A1%AB%E9%85%B8%E7%9B%90%E8%BF%98%E5%8E%9F%E8%8F%8C%E5%88%86%E7%A6%BB&cl=3&ie=utf-8&bs=%E5%90%91%E5%B1%B1%E5%9E%83%E5%9C%BE%E5%A1%AB%E5%9F%8B%E5%9C%BA-%E5%B0%BE%E7%9F%BF%E6%B7%B7%E5%90%88&f=8&rsv_bp=1&rsv_sug2=1&sc_f_para=sc_tasktype%3D%7BfirstSimpleSearch%7D&rsv_spt=3&rsv_n=2')
        html.encoding = 'utf-8'
        html = html.text
        # print(html)
        #如果名字不够精确，取结果页的第一个
        result = re.findall(r'ref_wr',html)   #[]为非详情页
        # print(result)
        if not result:
            #取结果页的第一个
            firstUrl = re.findall(r'<a href="(.*?)" data-click="{\'button_tp\':\'title\'}" target="_blank',html)[0]
            firstUrl = 'http://xueshu.baidu.com' + firstUrl
            # print(firstUrl)
            html = requests.get(firstUrl,headers=self.headers)
            html.encoding = 'utf-8'
            html = html.text
        return html
            #http://xueshu.baidu.com/s?wd=paperuri%3A%28c44e75d2af3457cf1ddb8da757e041bb%29&filter=sc_long_sign&sc_ks_para=q%3D%E5%90%91%E5%B1%B1%E5%9E%83%E5%9C%BE%E5%A1%AB%E5%9F%8B%E5%9C%BA-%E5%B0%BE%E7%9F%BF%E6%B7%B7%E5%90%88%E5%B8%A6%E7%A1%AB%E9%85%B8%E7%9B%90%E8%BF%98%E5%8E%9F%E8%8F%8C%E5%88%86%E7%A6%BB&sc_us=10296064545784877877&tn=SE_baiduxueshu_c1gjeupa&ie=utf-8
    def getNameList(self):
        try:
            nameFile = openpyxl.load_workbook(self.namexlsxPath)
            sheets = nameFile.sheetnames
            sheet0 = nameFile[sheets[0]]
            maxrow = sheet0.max_row
            papername = []
            celllist = sheet0['A']
            for cell in celllist:
                papername.append(cell.value)
            paperN = len(papername)
            print('共有%d篇文章\n' % paperN)
            self.paperNameList = papername
        except:
            print('读取论文名称表格出错！\n')

    def setQuoteData(self,html):
        quote = re.findall(r'class="sc_q" href="javascript:;" data-link="(.*?)" data-sign="', html)[0]
        diversion = re.findall(r'diversion="(.*?)"', html)[0]
        data_sign = re.findall(r'data-sign="(.*?)" data-click', html)[0]
        self.endnote_para['url'] = quote
        self.endnote_para['sign'] = data_sign
        self.endnote_para['diversion'] = diversion
        if not ( quote and diversion and data_sign ):
            print('下载链接获取失败！\n')


    def downQuote(self,paperName):
        file = requests.get(self.downmainurl, headers=self.headers, params=self.endnote_para)
        if len(paperName) > 230:
            paperName = paperName[0:100]
        # with方式，省略close
        try:
            with open(self.savePath + "/" + paperName + ".enw", 'wb') as f:
                f.write(file.content)
        except IOError:
            # print(len(paperName))
            print('写入文件错误！\n')
            pass

    def startDown(self):
        self.getNameList()
        for name in self.paperNameList:
            try:
                html = self.getMainHtml(name)
                self.setQuoteData(html)
                self.downQuote(name)
            except:
                print('出错题目：%s\n' % name)
                continue
    def setTimesleep(self,time):
        self.timesleep = time

    def setFormat(self,myformat):
        self.quote_format = myformat

